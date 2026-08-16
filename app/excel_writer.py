"""Write futures data into a formatted .xlsx workbook (openpyxl)."""

from __future__ import annotations

import datetime as _dt
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .downloader import FuturesPage

HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")  # light blue, black text
HEADER_FONT = Font(color="FF000000", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _write_rows(ws, start_row: int, headers: List[str], rows: List[list]) -> int:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BORDER
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.###"
        r += 1
    return r - 1


def _display_width(value) -> int:
    """Excel column width units: CJK chars render ~2 units, ASCII ~1."""
    w = 0
    for ch in str(value):
        w += 2 if ord(ch) > 0x2E7F else 1
    return w


def _autofit(ws, ncols: int, max_width: int = 44):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        width = 0
        for cell in ws[col_letter]:  # ws['C'] = column C (ws[3] would be row 3)
            if cell.value is None:
                continue
            width = max(width, _display_width(cell.value))
        ws.column_dimensions[col_letter].width = min(
            max(width + 3, 12), max_width
        )


def build_workbook(page: FuturesPage, title: str = "") -> Workbook:
    """Create a workbook with summary + detail sheets for one futures page."""
    wb = Workbook()
    ws = wb.active
    ws.title = "報價摘要"

    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A1"] = title or f"{page.contract_name or '期貨'} 報價"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"資料來源: https://www.etnet.com.hk/www/tc/futures/"
    ws["A3"] = f"網頁更新時間: {page.update_time or '-'}    下載時間: {now}"
    ws["A4"] = f"合約: {page.contract_name or ''}    (代碼: {page.subtype} {page.month})"

    start = 6
    headers = [
        "時段", "合約", "最新", "升跌", "升跌%", "高/低水",
        "最高", "最低", "前收市", "開市", "成交張數", "交易宗數", "每宗成交",
    ]
    rows = []
    for s in page.sessions:
        rows.append([
            s.session, s.contract, s.last, s.change, s.change_pct, s.premium,
            s.high, s.low, s.prev_close, s.open, s.volume, s.trades, s.avg_trade,
        ])
    _write_rows(ws, start, headers, rows)

    # --- 未平倉 sheet ---
    if page.open_interest is not None:
        ws2 = wb.create_sheet("未平倉")
        oi = page.open_interest
        _write_rows(
            ws2, 1,
            ["合約", "到期日", "未平倉總數 (GOI)", "未平倉淨數 (NOI)"],
            [[oi.contract or page.contract_name, oi.expiry, oi.goi, oi.noi]],
        )
        _autofit(ws2, 4)

    # --- 現貨 sheet ---
    if page.spot is not None:
        ws3 = wb.create_sheet("現貨")
        sp = page.spot
        _write_rows(
            ws3, 1,
            ["名稱", "最新", "升跌", "升跌%", "最高", "最低", "前收市", "開市"],
            [[sp.name, sp.last, sp.change, sp.change_pct,
              sp.high, sp.low, sp.prev_close, sp.open]],
        )
        _autofit(ws3, 8)

    # --- 15分鐘時段記錄 sheet ---
    if page.interval:
        ws4 = wb.create_sheet("15分鐘時段記錄")
        headers4 = [
            "時間", "開市價", "最高價", "最低價", "最新",
            "今日升跌", "升跌%", "高/低水", "成交", "交易宗數", "每宗成交",
        ]
        rows4 = [
            [r.time, r.open, r.high, r.low, r.last,
             r.change, r.change_pct, r.premium, r.volume, r.trades, r.avg_trade]
            for r in page.interval
        ]
        _write_rows(ws4, 1, headers4, rows4)
        _autofit(ws4, len(headers4))

    _autofit(ws, len(headers))
    return wb


def build_multi_workbook(
    pages: List[FuturesPage], title: str = "指數期貨報價"
) -> Workbook:
    """Workbook containing several contracts: one summary sheet + per-contract
    detail sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "總覽"
    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"資料來源: https://www.etnet.com.hk/www/tc/futures/    下載時間: {now}"

    headers = [
        "合約", "時段", "最新", "升跌", "升跌%", "高/低水",
        "最高", "最低", "前收市", "開市", "成交張數", "交易宗數", "每宗成交",
    ]
    rows = []
    for page in pages:
        for s in page.sessions:
            rows.append([
                page.contract_name, s.session, s.last, s.change, s.change_pct,
                s.premium, s.high, s.low, s.prev_close, s.open,
                s.volume, s.trades, s.avg_trade,
            ])
    _write_rows(ws, 4, headers, rows)
    _autofit(ws, len(headers))

    for page in pages:
        code = page.subtype or "UNKNOWN"
        # 報價 sheet
        ws_q = wb.create_sheet(f"報價_{code}")
        _write_rows(
            ws_q, 1,
            ["時段", "合約", "最新", "升跌", "升跌%", "高/低水",
             "最高", "最低", "前收市", "開市", "成交張數", "交易宗數", "每宗成交"],
            [[s.session, s.contract, s.last, s.change, s.change_pct, s.premium,
              s.high, s.low, s.prev_close, s.open, s.volume, s.trades, s.avg_trade]
             for s in page.sessions],
        )
        _autofit(ws_q, 13)

        if page.open_interest is not None:
            oi = page.open_interest
            ws_oi = wb.create_sheet(f"未平倉_{code}")
            _write_rows(
                ws_oi, 1,
                ["合約", "到期日", "未平倉總數 (GOI)", "未平倉淨數 (NOI)"],
                [[oi.contract or page.contract_name, oi.expiry, oi.goi, oi.noi]],
            )
            _autofit(ws_oi, 4)

        if page.interval:
            ws_i = wb.create_sheet(f"時段記錄_{code}")
            headers_i = [
                "時間", "開市價", "最高價", "最低價", "最新",
                "今日升跌", "升跌%", "高/低水", "成交", "交易宗數", "每宗成交",
            ]
            _write_rows(
                ws_i, 1, headers_i,
                [[r.time, r.open, r.high, r.low, r.last,
                  r.change, r.change_pct, r.premium, r.volume, r.trades,
                  r.avg_trade]
                 for r in page.interval],
            )
            _autofit(ws_i, len(headers_i))

        if page.spot is not None:
            sp = page.spot
            ws_sp = wb.create_sheet(f"現貨_{code}")
            _write_rows(
                ws_sp, 1,
                ["名稱", "最新", "升跌", "升跌%", "最高", "最低", "前收市", "開市"],
                [[sp.name, sp.last, sp.change, sp.change_pct,
                  sp.high, sp.low, sp.prev_close, sp.open]],
            )
            _autofit(ws_sp, 8)

    return wb


def default_filename(page: FuturesPage, prefix: str = "etnet_futures") -> str:
    """<prefix>_<code>_<month>_<yyyymmdd_hhmm>.xlsx"""
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    code = page.subtype or "ALL"
    month = page.month or ""
    return f"{prefix}_{code}_{month}_{stamp}.xlsx"


# ---------------------------------------------------------------------------
# Predefined per-product tabs (current + next month)
# ---------------------------------------------------------------------------

QUOTE_HEADERS = [
    "時段", "最新", "升跌", "升跌%", "高/低水",
    "最高", "最低", "前收市", "開市", "成交張數", "交易宗數", "每宗成交",
]
INTERVAL_HEADERS = [
    "時間", "開市價", "最高價", "最低價", "最新",
    "今日升跌", "升跌%", "高/低水", "成交", "交易宗數", "每宗成交",
]


def _section_title(ws, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="1F4E78")
    return row + 1


def build_tabs_workbook(pages_by_code: dict, product_names: dict = None,
                        title: str = "指數期貨報價") -> Workbook:
    """One sheet per product, each with 即月 + 下月 sections:
    quotes (日市/夜市), 未平倉, 15分鐘時段記錄."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for code, pages in pages_by_code.items():
        name = (product_names or {}).get(code, code)
        sheet_title = str(name)[:31]
        ws = wb.create_sheet(title=sheet_title)

        ws["A1"] = f"{name} ({code})"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"資料來源: https://www.etnet.com.hk/www/tc/futures/    下載時間: {now}"
        row = 4

        for page in pages:
            month_label = f"{page.month[:4]}/{page.month[4:]}" if len(page.month) == 6 else page.month
            row = _section_title(ws, row, f"{name} — {month_label}")
            # --- quotes ---
            _write_rows(ws, row, QUOTE_HEADERS, [
                [s.session, s.last, s.change, s.change_pct, s.premium,
                 s.high, s.low, s.prev_close, s.open,
                 s.volume, s.trades, s.avg_trade]
                for s in page.sessions
            ])
            row = row + 1 + max(len(page.sessions), 1)

            # --- open interest ---
            if page.open_interest is not None:
                oi = page.open_interest
                row = _section_title(ws, row, "未平倉")
                row = _write_rows(ws, row, ["合約", "到期日", "未平倉總數 (GOI)", "未平倉淨數 (NOI)"],
                                  [[oi.contract or name, oi.expiry, oi.goi, oi.noi]]) + 2

            # --- interval table (labels always shown, data when available) ---
            row = _section_title(ws, row, "15分鐘時段記錄")
            row = _write_rows(ws, row, INTERVAL_HEADERS, [
                [r.time, r.open, r.high, r.low, r.last,
                 r.change, r.change_pct, r.premium,
                 r.volume, r.trades, r.avg_trade]
                for r in page.interval
            ]) + 2

        _autofit(ws, max(len(QUOTE_HEADERS), len(INTERVAL_HEADERS)))

    return wb
